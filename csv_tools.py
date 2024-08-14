import os
import time

import pandas as pd
from gemini import gemini, prompt1,prompt2
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
def apply_function_and_update(df: pd.DataFrame, input_col: str, row_idx: int, func, output_col: str) -> None:
    """
    Apply a function to a value in a specific column and row of a DataFrame,
    and update the value in another specified column of the same row with the result.

    Parameters:
    df (pd.DataFrame): The DataFrame to operate on.
    input_col (str): The name of the column to get the input value from.
    row_idx (int): The index of the row to operate on.
    func (callable): The function to apply to the input value.
    output_col (str): The name of the column to update with the result.
    """
    # Get the value from the specified column and row
    value = df.at[row_idx, input_col]

    # Apply the function to the value
    result = func(value)

    # Update the specified column in the same row with the result
    df.at[row_idx, output_col] = result

# Example usage:
# Assuming you have a DataFrame 'df' and a function 'some_function'

# for row_index in range(len(df)):
#     apply_function_and_update(df, 'input_column', row_index, some_function, 'output_column')


import pandas as pd


def load_excel_to_dataframe(file_path):
    """
    Load an Excel file into a pandas DataFrame.

    Parameters:
    file_path (str): The path to the Excel file.

    Returns:
    pd.DataFrame: The loaded DataFrame.
    """
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        print("Excel file loaded successfully!")
        return df
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return None


import pandas as pd


def process_dataframe(df, column_number, newname, prompt1, result_df=None):
    # Initialize an empty list to store results
    results = []

    # Extract the text from the specified column
    for i in range(1, len(df)):
        text = str(df.iloc[i, column_number])
        result = gemini(prompt1, text)
        results.append(result)

    # Create a new DataFrame from the results list
    new_column_df = pd.DataFrame(results, columns=[newname])

    # If result_df is not provided, create a new DataFrame with the same index as df
    if result_df is None:
        result_df = pd.DataFrame(index=df.index)

    # Add the new column to the result DataFrame
    result_df = result_df.join(new_column_df)

    return result_df


def asis_df(df, column_number, newname, result_df=None):
    # Initialize an empty list to store results
    results = []

    # Extract the text from the specified column
    for i in range(1, len(df)):
        text = str(df.iloc[i, column_number])
        results.append(text)

    # Create a new DataFrame from the results list
    new_column_df = pd.DataFrame(results, columns=[newname])

    # If result_df is not provided, create a new DataFrame with the same index as df
    if result_df is None:
        result_df = pd.DataFrame(index=df.index)

    # Add the new column to the result DataFrame
    result_df = result_df.join(new_column_df)

    return result_df

def save_dataframe(df, format, rtl=True):
    directory = 'media/download'
    if not os.path.exists(directory):
        os.makedirs(directory)

    filename = os.path.join(directory, 'Updated Table')

    if format.lower() == 'csv':
        full_filename = filename + '.csv'
        df.to_csv(full_filename, index=False)
        print(f"DataFrame saved as {full_filename}")
    elif format.lower() == 'xlsx':
        full_filename = filename + '.xlsx'
        df.to_excel(full_filename, index=False)
        if rtl:
            # Load the workbook and select the active sheet
            wb = load_workbook(full_filename)
            ws = wb.active

            # Set the sheet to RTL orientation
            ws.sheet_view.rightToLeft = True

            # Save the modified workbook with RTL support
            wb.save(full_filename)
            print(f"DataFrame saved as {full_filename} with RTL orientation")
        else:
            print(f"DataFrame saved as {full_filename}")
    else:
        raise ValueError("Unsupported format. Please choose either 'csv' or 'xlsx'.")


# Example usage
# df = load_excel_to_dataframe('Test - Copy.xlsx')
# new_df= asis_df(df, 0, "כותרת")
# new_df = process_dataframe(df, 6, 'תקציר על יזם', prompt1, new_df)
# new_df = process_dataframe(df, 16, 'תקציר על האתגר', prompt2, new_df)
# save_dataframe(new_df, 'csv')
# save_dataframe(new_df, 'xlsx')
import pandas as pd
import time


def score_Bashlot(df):
    results = []
    total_rows = len(df) - 1

    prompt_start_date = """
    the next text is a date - return me the score which is how long ago was it:
    1 - up to 6 months
    2 - 6-12 months
    3 - 1-2 years
    4 - 2-4 years
    5 - 4+ years
    return only the score as a number
    """

    prompt_people_experienced = """
    For the next text return me the score which is how many people tried the service:
    1 - 0-2 
    2 - 3-12 
    3 - 13-100 
    4 - 101-1000 
    5 - 1000 and above 
    return only the score as a number
    """

    prompt_salaried_employees = """
    the next text will be contain a value which is a number in int or str
    please return the following score only 
    1 score for value 0 
    2 score for value 1
    3 score for value in interval [2-3]
    4 score for value in interval [4-10]
    5 score for value 11 and above

    return score only
    """

    prompt_funds_raised = """
    For the next text return me the score which is how much budget the project has:
    1 - 0
    2 - up to 30,000
    3 - 30,001-100,000
    4 - 101,000-500,000
    5 - 500,000 and above
    return only the score as a number
    """

    prompt_text_analysis = """
    For the next text return me the score which is how the precentge of this hole words:
    פיילוט, התנסות, פידבקים, תוצאות, משתמשים, שותפים, שירות, התמדה, נשירה, design partner, retention rate, לקוחות, משלמים, פעילים/ פעילות
    exist in the next text :
    1 - 0-20%
    2 - 21-40%
    3 - 41-60%
    4 - 61-80%
    5 - 81-100%
    return only the score as a number
    """

    for i in range(len(df)):
        text1 = str(df.iloc[i, 1])
        date = int(gemini(prompt_start_date, text1))
        print(f'date : {date}')

        text2 = str(df.iloc[i, 5])
        tried = int(gemini(prompt_people_experienced, text2))
        print(f'tried : {tried}')
        time.sleep(3)

        text3 = str(df.iloc[i, 8])
        employ = int(gemini(prompt_salaried_employees, text3))
        print(f'employ : {employ}')

        text4 = str(df.iloc[i, 9])
        money = int(gemini(prompt_funds_raised, text4))
        print(f'money : {money}')

        text5 = str(df.iloc[i, 11] + df.iloc[i, 12] + df.iloc[i, 14] + df.iloc[i, 15] + df.iloc[i, 16])
        percentage = int(gemini(prompt_text_analysis, text5))
        print(f'percentage : {percentage}')
        time.sleep(3)

        result = 0.1 * date + 0.3 * tried + 0.1 * employ + 0.15 * money + 0.25 * percentage
        formatted_result = f"{result:.2f}"
        print(f'result : {formatted_result}')

        results.append(formatted_result)

    new_column_df = pd.DataFrame(results, columns=["בשלות"])
    result_df = df.join(new_column_df)
    return result_df


prompt1 = "קח את הטקסט הבא בעברית וסכם את הרקע של היזם/יזמים למשפט אחד."
prompt2 = "קח את הטקסט הבא בעברית וסכם את האתגר הספציפי שהמיזם מנסה לפתור בשני משפטים."
prompt3 = "קח את הטקסט הבא בעברית ותאר כיצד המיזם נותן מענה לאתגר הספציפי בשני משפטים."
prompt4 = "קח את הסיכומים הבאים וכתוב סיכום חדש על המיזם באורך של עד 75 מילים. תן דגש לקראת הסוף לדרך הפעולה במיזם"
prompt5 = "קח את הטקסט הבא בעברית וסכם מה נעשה עד היום במיזם למשפט אחד. אם כתוב טקסט ללא הרבה מידע - כתוב שלא נעשה עד היום יותר מידי"
prompt6 = "קח את הסיכומים הבאים וכתוב סיכום חדש על המיזם באורך למשפט אחד בעברית"



def process_dataframe_summary(df, result_df=None):
    global progress
    results = []

    for i in range(len(df)):
        text = str(df.iloc[i,10])
        print(text)
        answer1 = gemini(prompt2, text)
        text2 = str(df.iloc[i,11])
        answer2 = gemini(prompt3, text2)
        print(answer2)

        time.sleep(3)

        text3 = str(df.iloc[i,15])
        answer3 = gemini(prompt5, text3)
        print(answer3)

        text4 = " הסיכומים הינם " + answer1 + answer2 + answer3
        answer4 = gemini(prompt4, text4)
        print(answer4)
        results.append(answer4)
        time.sleep(5)

        # Update progress
    newname = "סיכום בפסקה"

    new_column_df = pd.DataFrame(results, columns=[newname])

    if result_df is None:
        result_df = pd.DataFrame(index=df.index)

    result_df = result_df.join(new_column_df)
    return result_df


def process_dataframe_summary_one_line(df, column_number, newname, prompt1, result_df=None):
    global progress
    results = []

    for i in range(len(df)):
        text = str(df.iloc[i, 10])
        print(text)
        answer1 = gemini(prompt2, text)
        time.sleep(3)
        text2 = str(df.iloc[i, 11])
        answer2 = gemini(prompt3, text2)
        time.sleep(3)

        text3 = str(df.iloc[i, 15])
        answer3 = gemini(prompt5, text3)
        time.sleep(3)

        text4 = answer1 + answer2 + answer3
        answer4 = gemini(prompt6, text4)
        results.append(answer4)
        time.sleep(3)

        # Update progress
    newname = "סיכום בשורה"
    new_column_df = pd.DataFrame(results, columns=[newname])

    if result_df is None:
        result_df = pd.DataFrame(index=df.index)

    result_df = result_df.join(new_column_df)
    return result_df

# # Display the resulting DataFrame
# print(new_df)
#
# new_df.to_excel('new_file.xlsx', index=False)
# # Load the workbook and select the active sheet
# wb = load_workbook('new_file.xlsx')
# ws = wb.active
#
# # Set the sheet to RTL orientation
# ws.sheet_view.rightToLeft = True
#
# # Save the modified workbook
# wb.save('new_file_rtl.xlsx')


df = load_excel_to_dataframe('book1.xlsx')
a = process_dataframe_summary(df, df)
print(a)
