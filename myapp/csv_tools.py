import os

import pandas as pd
from myapp.gemini import gemini, prompt1,prompt2
from openpyxl import load_workbook

# Example usage:
# Assuming you have a DataFrame 'df' and a function 'some_function'

# for row_index in range(len(df)):
#     apply_function_and_update(df, 'input_column', row_index, some_function, 'output_column')


def load_excel_to_dataframe(file_path):
    """
    Load an Excel file into a pandas DataFrame.

    Parameters:
    file_path (str): The path to the Excel file.

    Returns:
    pd.DataFrame: The loaded DataFrame.
    """
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        print("Excel file loaded successfully!")
        return df
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return None


import pandas as pd


def process_dataframe(df, column_number, newname, prompt1, result_df=None):
    # Initialize an empty list to store results
    results = []

    # Extract the text from the specified column
    for i in range(1, len(df)):
        text = str(df.iloc[i, column_number])
        result = gemini(prompt1, text)
        results.append(result)

    # Create a new DataFrame from the results list
    new_column_df = pd.DataFrame(results, columns=[newname])

    # If result_df is not provided, create a new DataFrame with the same index as df
    if result_df is None:
        result_df = pd.DataFrame(index=df.index)

    # Add the new column to the result DataFrame
    result_df = result_df.join(new_column_df)

    return result_df


def asis_df(df, column_number, newname, result_df=None):
    # Initialize an empty list to store results
    results = []

    # Extract the text from the specified column
    for i in range(1, len(df)):
        text = str(df.iloc[i, column_number])
        results.append(text)

    # Create a new DataFrame from the results list
    new_column_df = pd.DataFrame(results, columns=[newname])

    # If result_df is not provided, create a new DataFrame with the same index as df
    if result_df is None:
        result_df = pd.DataFrame(index=df.index)

    # Add the new column to the result DataFrame
    result_df = result_df.join(new_column_df)

    return result_df

def save_dataframe(df, format, rtl=True):
    directory = 'media/download'
    if not os.path.exists(directory):
        os.makedirs(directory)

    filename = os.path.join(directory, 'Updated Table')

    if format.lower() == 'csv':
        full_filename = filename + '.csv'
        df.to_csv(full_filename, index=False)
        print(f"DataFrame saved as {full_filename}")
    elif format.lower() == 'xlsx':
        full_filename = filename + '.xlsx'
        df.to_excel(full_filename, index=False)
        if rtl:
            # Load the workbook and select the active sheet
            wb = load_workbook(full_filename)
            ws = wb.active

            # Set the sheet to RTL orientation
            ws.sheet_view.rightToLeft = True

            # Save the modified workbook with RTL support
            wb.save(full_filename)
            print(f"DataFrame saved as {full_filename} with RTL orientation")
        else:
            print(f"DataFrame saved as {full_filename}")
    else:
        raise ValueError("Unsupported format. Please choose either 'csv' or 'xlsx'.")


# Example usage
df = load_excel_to_dataframe('Test - Copy.xlsx')
new_df= asis_df(df, 0, "כותרת")
new_df = process_dataframe(df, 6, 'תקציר על יזם', prompt1, new_df)
new_df = process_dataframe(df, 16, 'תקציר על האתגר', prompt2, new_df)
save_dataframe(new_df, 'xlsx')
save_dataframe(new_df, 'csv')

# # Display the resulting DataFrame
# print(new_df)
#
# new_df.to_excel('new_file.xlsx', index=False)
# # Load the workbook and select the active sheet
# wb = load_workbook('new_file.xlsx')
# ws = wb.active
#
# # Set the sheet to RTL orientation
# ws.sheet_view.rightToLeft = True
#
# # Save the modified workbook
# wb.save('new_file_rtl.xlsx')


